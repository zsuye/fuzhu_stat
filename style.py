from openpyxl import load_workbook
from typing import Dict, Any


def read_excel_styles(file_path: str, sheet_name: str = None) -> Dict[str, Any]:
    """
    读取Excel文件的第一行和第二行的样式信息（包括行高、列宽、居中和粗体）。

    参数:
        file_path (str): Excel文件的路径。
        sheet_name (str, 可选): 工作表名。如果为None，则使用激活的工作表。

    返回:
        Dict[str, Any]: 包含样式信息的字典。
    """
    # 加载Excel文件
    wb = load_workbook(file_path)
    ws = wb[sheet_name] if sheet_name else wb.active

    # 初始化样式信息字典
    styles_info = {
        "first_row": {"height": None, "styles": {}},  # 存储各列的样式
        "second_row": {"height": None, "styles": {}},  # 存储各列的样式
    }

    # 获取第一行和第二行的行高
    styles_info["first_row"]["height"] = ws.row_dimensions[1].height
    styles_info["second_row"]["height"] = ws.row_dimensions[2].height

    # 获取第一行和第二行的各列样式
    for row_num in [1, 2]:
        row_key = "first_row" if row_num == 1 else "second_row"
        for cell in ws[row_num]:
            col_letter = cell.column_letter
            styles_info[row_key]["styles"][col_letter] = {
                "width": ws.column_dimensions[col_letter].width,
                "bold": cell.font.bold,
                "alignment": cell.alignment.horizontal,  # 'center', 'left', 'right' 等
            }

    return styles_info


# 由于当前环境中无法访问文件，您可以在本地环境中运行这个函数来获取样式信息。
# 例如：
# styles_info = read_excel_styles("E:/桂香园/10月31日电商发货表.xlsx", "去快递费单价")
# print(styles_info)

sheet1_style = {
    "first_row": {
        "height": 18.5,
        "styles": {
            "A": {"width": 12.3416666666667, "bold": True, "alignment": "center"},
            "B": {"width": 23.675, "bold": True, "alignment": "center"},
            "C": {"width": 11.3416666666667, "bold": True, "alignment": "center"},
            "D": {"width": 8.50833333333333, "bold": True, "alignment": "center"},
            "E": {"width": 14.675, "bold": True, "alignment": "center"},
            "F": {"width": 10.3416666666667, "bold": True, "alignment": "center"},
            "G": {"width": 19.0083333333333, "bold": True, "alignment": "center"},
            "H": {"width": 10.3416666666667, "bold": True, "alignment": "center"},
            "I": {"width": 13.0, "bold": True, "alignment": "center"},
            "J": {"width": 8.34166666666667, "bold": True, "alignment": "center"},
            "K": {"width": 14.8416666666667, "bold": True, "alignment": "center"},
            "L": {"width": 12.175, "bold": True, "alignment": "center"},
            "M": {"width": 12.675, "bold": True, "alignment": "center"},
            "N": {"width": 12.175, "bold": True, "alignment": "center"},
        },
    },
    "second_row": {
        "height": 18.5,
        "styles": {
            "A": {"width": 12.3416666666667, "bold": False, "alignment": "center"},
            "B": {"width": 23.675, "bold": False, "alignment": "center"},
            "C": {"width": 11.3416666666667, "bold": False, "alignment": "center"},
            "D": {"width": 8.50833333333333, "bold": False, "alignment": "center"},
            "E": {"width": 14.675, "bold": False, "alignment": "center"},
            "F": {"width": 10.3416666666667, "bold": False, "alignment": "center"},
            "G": {"width": 19.0083333333333, "bold": False, "alignment": "center"},
            "H": {"width": 10.3416666666667, "bold": False, "alignment": "center"},
            "I": {"width": 13.0, "bold": False, "alignment": "center"},
            "J": {"width": 8.34166666666667, "bold": False, "alignment": "center"},
            "K": {"width": 14.8416666666667, "bold": False, "alignment": "center"},
            "L": {"width": 12.175, "bold": False, "alignment": "center"},
            "M": {"width": 12.675, "bold": False, "alignment": "center"},
            "N": {"width": 12.175, "bold": False, "alignment": "center"},
        },
    },
}

sheet2_style = {
    "first_row": {
        "height": 22.5,
        "styles": {
            "A": {"width": 16.8416666666667, "bold": True, "alignment": "center"},
            "B": {"width": 16.0083333333333, "bold": True, "alignment": "center"},
        },
    },
    "second_row": {
        "height": 22.5,
        "styles": {
            "A": {"width": 16.8416666666667, "bold": False, "alignment": "center"},
            "B": {"width": 16.0083333333333, "bold": False, "alignment": "center"},
        },
    },
}
